import datetime
import calendar
import re
from dataclasses import dataclass, field
from openpyxl import load_workbook
from config import RE_INVOICE, RE_CHECK, RE_CREDIT_CARD, RE_WIRE, RE_BALANCE_FORWARD, SUMMARY_KEYWORDS

# "Fatura" in the description with no SH-prefixed invoice number usually means
# a credit memo / refund (e.g. "Fatura No Sel202600000092"). When combined with
# an amount in the credit column, it should settle outstanding AR rather than
# be treated as a zero-amount invoice.
RE_FATURA_ANY = re.compile(r'Fatura\s+No\s+\w+', re.IGNORECASE)


@dataclass
class Transaction:
    row_index: int
    amount_debit: float       # Column A — what customer owes (invoices)
    amount_credit: float      # Column B — what customer paid (payments)
    description: str          # Column C
    notes: str                # Column D
    payment_terms: str        # Column E — Vadeli/KK/Pesin
    date: datetime.date       # Column F
    invoice_number: str       # Column G
    running_balance: float    # Column H
    tx_type: str              # invoice / check / credit_card / wire / balance_forward
    parsed_details: dict = field(default_factory=dict)


@dataclass
class StatementData:
    customer_name: str
    period_start: datetime.date
    period_end: datetime.date
    transactions: list
    total_debit: float
    total_credit: float
    end_balance: float


def correct_settlement_date(day, month, year):
    """Roll invalid dates (e.g. 31st in a 30-day month) to 1st of next month."""
    max_day = calendar.monthrange(year, month)[1]
    if day > max_day:
        if month == 12:
            return datetime.date(year + 1, 1, 1)
        return datetime.date(year, month + 1, 1)
    return datetime.date(year, month, day)


def parse_check_date(date_str):
    """Parse DD/MM/YYYY string with month-end correction."""
    parts = date_str.split("/")
    day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
    return correct_settlement_date(day, month, year)


def classify_transaction(description):
    """Classify transaction type and extract details from description."""
    if not description:
        return "unknown", {}

    desc = str(description).strip()

    m = RE_BALANCE_FORWARD.search(desc)
    if m:
        return "balance_forward", {}

    m = RE_INVOICE.search(desc)
    if m:
        return "invoice", {"invoice_number": m.group(1)}

    m = RE_CREDIT_CARD.search(desc)
    if m:
        return "credit_card", {"receipt_number": m.group(1)}

    m = RE_WIRE.search(desc)
    if m:
        return "wire", {}

    m = RE_CHECK.search(desc)
    if m:
        settlement_date = parse_check_date(m.group(2))
        return "check", {
            "payee_name": m.group(1).strip(),
            "settlement_date": settlement_date,
            "bank": m.group(3).strip(),
            "check_number": m.group(4),
        }

    return "unknown", {}


def to_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return 0.0


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def is_summary_row(description):
    """Check if a row is a total/balance summary row."""
    if not description:
        return False
    desc = str(description)
    return any(kw in desc for kw in SUMMARY_KEYWORDS)


def parse_excel(file_obj):
    """Parse an Excel balance statement and return StatementData."""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    customer_name = str(ws["B1"].value or "Unknown")
    period_start = to_date(ws["G1"].value)
    period_end = to_date(ws["I1"].value)

    transactions = []
    total_debit = 0.0
    total_credit = 0.0
    end_balance = 0.0

    # Scan from the bottom to find where data ends and summary rows begin.
    # Summary rows are identified by keywords in Column C (مجموع, الاجمالي,
    # الرصيد, Toplam, etc.) rather than a fixed offset, because different
    # templates have a variable number of summary rows.
    max_row = ws.max_row
    summary_rows = []
    row = max_row
    # Skip trailing fully-blank rows (A=0/None, B=0/None, C=None). Some
    # exports include an empty sentinel row after the totals that would
    # otherwise halt the scan before we find the real summary rows.
    while row >= 3:
        a = to_float(ws.cell(row=row, column=1).value)
        b = to_float(ws.cell(row=row, column=2).value)
        c = ws.cell(row=row, column=3).value
        if a == 0 and b == 0 and not c:
            row -= 1
        else:
            break
    while row >= 3:
        desc = ws.cell(row=row, column=3).value
        if is_summary_row(desc):
            summary_rows.append(row)
            row -= 1
        else:
            break

    # The last data row is the one above all summary rows.
    data_end_row = row

    # Identify the totals row (contains both A and B amounts) and the
    # balance row (contains only one amount — the end balance).
    total_debit = 0.0
    total_credit = 0.0
    end_balance = 0.0
    for sr in sorted(summary_rows):
        a = to_float(ws.cell(row=sr, column=1).value)
        b = to_float(ws.cell(row=sr, column=2).value)
        if a > 0 and b > 0:
            # Totals row — keep the largest we find (handles dup grand-total rows)
            if a > total_debit:
                total_debit = a
                total_credit = b
        else:
            # Balance row — only one column has a value
            end_balance = a if a != 0 else b

    for row_idx in range(3, data_end_row + 1):
        amount_debit = to_float(ws.cell(row=row_idx, column=1).value)
        amount_credit = to_float(ws.cell(row=row_idx, column=2).value)
        description = str(ws.cell(row=row_idx, column=3).value or "")
        notes = str(ws.cell(row=row_idx, column=4).value or "")
        payment_terms = str(ws.cell(row=row_idx, column=5).value or "")
        tx_date = to_date(ws.cell(row=row_idx, column=6).value)
        invoice_number = str(ws.cell(row=row_idx, column=7).value or "")
        running_balance = to_float(ws.cell(row=row_idx, column=8).value)

        if amount_debit == 0 and amount_credit == 0:
            continue

        tx_type, parsed_details = classify_transaction(description)

        # A row whose description looks like a "Fatura" entry but whose amount
        # sits in the credit column is a credit note / refund — money flowing
        # back to the customer. This covers both SH-prefixed invoices (which
        # classify as "invoice") and other prefixes like "Sel..." (which fall
        # through to "unknown" because RE_INVOICE is SH-only). Treat as a
        # payment-like entry so it reduces AR in FIFO matching rather than
        # being dropped as a zero-amount invoice.
        if (amount_debit == 0 and amount_credit > 0
                and (tx_type == "invoice" or RE_FATURA_ANY.search(description))):
            tx_type = "credit_note"
            parsed_details = {}

        transactions.append(Transaction(
            row_index=row_idx,
            amount_debit=amount_debit,
            amount_credit=amount_credit,
            description=description,
            notes=notes,
            payment_terms=payment_terms,
            date=tx_date,
            invoice_number=invoice_number,
            running_balance=running_balance,
            tx_type=tx_type,
            parsed_details=parsed_details,
        ))

    wb.close()

    return StatementData(
        customer_name=customer_name,
        period_start=period_start,
        period_end=period_end,
        transactions=transactions,
        total_debit=total_debit,
        total_credit=total_credit,
        end_balance=end_balance,
    )
